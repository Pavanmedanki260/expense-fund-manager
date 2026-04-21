import os
import uuid
import asyncio
import logging
import secrets
import bcrypt
import jwt as pyjwt
import resend
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Optional

from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, HTTPException, Request, Response, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

app = FastAPI(title="FundTrack API")

MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME")
JWT_SECRET = os.environ.get("JWT_SECRET")
JWT_ALGORITHM = "HS256"
RESEND_API_KEY = os.environ.get("RESEND_API_KEY")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "onboarding@resend.dev")
APP_URL = os.environ.get("APP_URL", "")
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@fundtrack.com")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "Admin@123")

resend.api_key = RESEND_API_KEY

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        APP_URL,
        "http://localhost:3000",
        "https://b844d7d4-15d5-4ca4-8fbd-5ff90237ce2f.preview.emergentagent.com",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]


# ── Password & JWT Helpers ──────────────────────────────────────
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(hours=24), "type": "access"}
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, access_token: str, refresh_token: str):
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")


# ── Models ──────────────────────────────────────────────────────
class RegisterBody(BaseModel):
    name: str
    email: str
    password: str

class LoginBody(BaseModel):
    email: str
    password: str

class GroupCreate(BaseModel):
    name: str
    description: Optional[str] = ""

class InviteCreate(BaseModel):
    email: str
    role: str

class FundCreate(BaseModel):
    source_name: str
    amount_inr: float
    category: str
    date_received: str
    notes: Optional[str] = ""
    attachment_url: Optional[str] = ""

class UtilizationCreate(BaseModel):
    purpose: str
    amount_inr: float
    date_spent: str
    linked_fund_id: str
    notes: Optional[str] = ""
    receipt_url: Optional[str] = ""

class MemberRoleUpdate(BaseModel):
    role: str


# ── Auth Helpers ────────────────────────────────────────────────
async def get_current_user(request: Request):
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"user_id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        if not user.get("is_active", True):
            raise HTTPException(status_code=403, detail="Account deactivated")
        user.pop("password_hash", None)
        return user
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except pyjwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


async def get_group_member_role(user_id: str, group_id: str):
    member = await db.group_members.find_one({"group_id": group_id, "user_id": user_id}, {"_id": 0})
    return member["role"] if member else None


async def require_group_role(request: Request, group_id: str, allowed_roles: list):
    user = await get_current_user(request)
    if user.get("is_super_admin"):
        return user, "super_admin"
    role = await get_group_member_role(user["user_id"], group_id)
    if not role:
        raise HTTPException(status_code=403, detail="Not a member of this group")
    if role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Insufficient permissions in this group")
    return user, role


def user_response(user: dict) -> dict:
    """Strip password_hash from user before returning."""
    u = {k: v for k, v in user.items() if k != "password_hash"}
    return u


# ── Startup ─────────────────────────────────────────────────────
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    # Seed admin
    existing = await db.users.find_one({"email": ADMIN_EMAIL.lower()}, {"_id": 0})
    if not existing:
        admin_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": admin_id,
            "name": "Admin",
            "email": ADMIN_EMAIL.lower(),
            "password_hash": hash_password(ADMIN_PASSWORD),
            "is_super_admin": True,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info(f"Admin seeded: {ADMIN_EMAIL}")
    else:
        if not verify_password(ADMIN_PASSWORD, existing.get("password_hash", "")):
            await db.users.update_one({"email": ADMIN_EMAIL.lower()}, {"$set": {"password_hash": hash_password(ADMIN_PASSWORD)}})
            logger.info("Admin password updated from env")


# ── Auth Endpoints ──────────────────────────────────────────────
@app.post("/api/auth/register")
async def register(body: RegisterBody, response: Response):
    email = body.email.strip().lower()
    if not body.name.strip():
        raise HTTPException(status_code=400, detail="Name is required")
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    await db.users.insert_one({
        "user_id": user_id,
        "name": body.name.strip(),
        "email": email,
        "password_hash": hash_password(body.password),
        "is_super_admin": False,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    # Auto-accept pending invitations
    pending = await db.invitations.find({"email": email, "status": "pending"}, {"_id": 0}).to_list(100)
    for inv in pending:
        exists = await db.group_members.find_one({"group_id": inv["group_id"], "user_id": user_id}, {"_id": 0})
        if not exists:
            await db.group_members.insert_one({"group_id": inv["group_id"], "user_id": user_id, "role": inv["role"], "joined_at": datetime.now(timezone.utc).isoformat()})
        await db.invitations.update_one({"invite_id": inv["invite_id"]}, {"$set": {"status": "accepted", "accepted_by_user_id": user_id}})

    access = create_access_token(user_id, email)
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return user_response(user)


@app.post("/api/auth/login")
async def login(body: LoginBody, request: Request, response: Response):
    email = body.email.strip().lower()
    ip = request.client.host if request.client else "unknown"
    ident = f"{ip}:{email}"
    # Brute force check
    attempt = await db.login_attempts.find_one({"identifier": ident}, {"_id": 0})
    if attempt and attempt.get("count", 0) >= 5:
        locked_until = attempt.get("locked_until")
        if locked_until:
            if isinstance(locked_until, str):
                locked_until = datetime.fromisoformat(locked_until)
            if locked_until.tzinfo is None:
                locked_until = locked_until.replace(tzinfo=timezone.utc)
            if datetime.now(timezone.utc) < locked_until:
                raise HTTPException(status_code=429, detail="Too many failed attempts. Try again in 15 minutes.")
            else:
                await db.login_attempts.delete_one({"identifier": ident})

    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not verify_password(body.password, user.get("password_hash", "")):
        # Increment failed attempts
        await db.login_attempts.update_one(
            {"identifier": ident},
            {"$inc": {"count": 1}, "$set": {"locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()}},
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")

    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Account deactivated")

    # Clear failed attempts
    await db.login_attempts.delete_one({"identifier": ident})

    access = create_access_token(user["user_id"], email)
    refresh = create_refresh_token(user["user_id"])
    set_auth_cookies(response, access, refresh)
    return user_response(user)


@app.get("/api/auth/me")
async def get_me(user=Depends(get_current_user)):
    return user


@app.post("/api/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/", samesite="none", secure=True)
    response.delete_cookie("refresh_token", path="/", samesite="none", secure=True)
    return {"message": "Logged out"}


@app.post("/api/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"user_id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        new_access = create_access_token(user["user_id"], user["email"])
        response.set_cookie(key="access_token", value=new_access, httponly=True, secure=True, samesite="none", max_age=86400, path="/")
        return {"message": "Token refreshed"}
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired")
    except pyjwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")


# ── Group Endpoints ─────────────────────────────────────────────
@app.post("/api/groups")
async def create_group(body: GroupCreate, user=Depends(get_current_user)):
    group_id = f"grp_{uuid.uuid4().hex[:12]}"
    await db.groups.insert_one({
        "group_id": group_id, "name": body.name, "description": body.description,
        "created_by_user_id": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat(),
    })
    await db.group_members.insert_one({"group_id": group_id, "user_id": user["user_id"], "role": "admin", "joined_at": datetime.now(timezone.utc).isoformat()})
    return await db.groups.find_one({"group_id": group_id}, {"_id": 0})


@app.get("/api/groups")
async def list_groups(user=Depends(get_current_user)):
    if user.get("is_super_admin"):
        groups = await db.groups.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    else:
        memberships = await db.group_members.find({"user_id": user["user_id"]}, {"_id": 0}).to_list(500)
        gids = [m["group_id"] for m in memberships]
        if not gids:
            return []
        groups = await db.groups.find({"group_id": {"$in": gids}}, {"_id": 0}).sort("created_at", -1).to_list(500)
    for g in groups:
        g["member_count"] = await db.group_members.count_documents({"group_id": g["group_id"]})
        if user.get("is_super_admin"):
            g["user_role"] = "super_admin"
        else:
            m = await db.group_members.find_one({"group_id": g["group_id"], "user_id": user["user_id"]}, {"_id": 0})
            g["user_role"] = m["role"] if m else "none"
    return groups


@app.get("/api/groups/{group_id}")
async def get_group(group_id: str, request: Request):
    user = await get_current_user(request)
    group = await db.groups.find_one({"group_id": group_id}, {"_id": 0})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if not user.get("is_super_admin"):
        role = await get_group_member_role(user["user_id"], group_id)
        if not role:
            raise HTTPException(status_code=403, detail="Not a member of this group")
        group["user_role"] = role
    else:
        group["user_role"] = "super_admin"
    group["member_count"] = await db.group_members.count_documents({"group_id": group_id})
    return group


@app.put("/api/groups/{group_id}")
async def update_group(group_id: str, body: GroupCreate, request: Request):
    user, role = await require_group_role(request, group_id, ["admin"])
    await db.groups.update_one({"group_id": group_id}, {"$set": {"name": body.name, "description": body.description, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return await db.groups.find_one({"group_id": group_id}, {"_id": 0})


@app.delete("/api/groups/{group_id}")
async def delete_group(group_id: str, request: Request):
    await require_group_role(request, group_id, ["admin"])
    await db.groups.delete_one({"group_id": group_id})
    await db.group_members.delete_many({"group_id": group_id})
    await db.funds.delete_many({"group_id": group_id})
    await db.utilizations.delete_many({"group_id": group_id})
    await db.invitations.delete_many({"group_id": group_id})
    return {"message": "Group deleted"}


# ── Group Members ───────────────────────────────────────────────
@app.get("/api/groups/{group_id}/members")
async def list_members(group_id: str, request: Request):
    user = await get_current_user(request)
    if not user.get("is_super_admin"):
        r = await get_group_member_role(user["user_id"], group_id)
        if not r:
            raise HTTPException(status_code=403, detail="Not a member")
    members = await db.group_members.find({"group_id": group_id}, {"_id": 0}).to_list(500)
    for m in members:
        u = await db.users.find_one({"user_id": m["user_id"]}, {"_id": 0, "password_hash": 0})
        if u:
            m["name"] = u.get("name", ""); m["email"] = u.get("email", ""); m["is_active"] = u.get("is_active", True)
    return members


@app.put("/api/groups/{group_id}/members/{member_user_id}/role")
async def update_member_role(group_id: str, member_user_id: str, body: MemberRoleUpdate, request: Request):
    await require_group_role(request, group_id, ["admin"])
    if body.role not in ["admin", "contributor", "viewer"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    existing = await db.group_members.find_one({"group_id": group_id, "user_id": member_user_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Member not found")
    await db.group_members.update_one({"group_id": group_id, "user_id": member_user_id}, {"$set": {"role": body.role}})
    return {"message": "Role updated"}


@app.delete("/api/groups/{group_id}/members/{member_user_id}")
async def remove_member(group_id: str, member_user_id: str, request: Request):
    await require_group_role(request, group_id, ["admin"])
    await db.group_members.delete_one({"group_id": group_id, "user_id": member_user_id})
    return {"message": "Member removed"}


# ── Invitations ─────────────────────────────────────────────────
@app.post("/api/groups/{group_id}/invite")
async def invite_member(group_id: str, body: InviteCreate, request: Request):
    user, role = await require_group_role(request, group_id, ["admin"])
    if body.role not in ["admin", "contributor", "viewer"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    group = await db.groups.find_one({"group_id": group_id}, {"_id": 0})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    inv_email = body.email.strip().lower()
    existing_user = await db.users.find_one({"email": inv_email}, {"_id": 0})
    if existing_user:
        em = await db.group_members.find_one({"group_id": group_id, "user_id": existing_user["user_id"]}, {"_id": 0})
        if em:
            raise HTTPException(status_code=400, detail="User is already a member")
    existing_inv = await db.invitations.find_one({"group_id": group_id, "email": inv_email, "status": "pending"}, {"_id": 0})
    if existing_inv:
        raise HTTPException(status_code=400, detail="Invitation already sent")

    invite_id = f"inv_{uuid.uuid4().hex[:12]}"
    await db.invitations.insert_one({
        "invite_id": invite_id, "group_id": group_id, "email": inv_email, "role": body.role,
        "invited_by_user_id": user["user_id"], "invited_by_name": user["name"],
        "status": "pending", "created_at": datetime.now(timezone.utc).isoformat(),
    })
    invite_link = f"{APP_URL}/invite/{invite_id}"
    html = f"""
    <div style="font-family:sans-serif;max-width:520px;margin:0 auto;padding:32px;">
      <div style="background:#1D9E75;color:white;padding:16px 24px;border-radius:12px 12px 0 0;text-align:center;"><h1 style="margin:0;font-size:22px;">FundTrack</h1></div>
      <div style="background:#fff;border:1px solid #E4E4E7;border-top:none;padding:32px 24px;border-radius:0 0 12px 12px;">
        <p style="font-size:16px;color:#374151;">Hi,</p>
        <p style="font-size:15px;color:#374151;"><strong>{user["name"]}</strong> invited you to join <strong>{group["name"]}</strong> as <strong style="color:#1D9E75;">{body.role}</strong>.</p>
        <div style="text-align:center;margin:28px 0;"><a href="{invite_link}" style="background:#1D9E75;color:white;padding:12px 32px;border-radius:8px;text-decoration:none;font-size:15px;font-weight:600;">Accept Invitation</a></div>
        <p style="font-size:13px;color:#6B7280;">Or copy: <a href="{invite_link}" style="color:#1D9E75;">{invite_link}</a></p>
      </div>
    </div>"""
    try:
        await asyncio.to_thread(resend.Emails.send, {"from": SENDER_EMAIL, "to": [inv_email], "subject": f"You're invited to join {group['name']} on FundTrack", "html": html})
    except Exception as e:
        logger.error(f"Email send failed: {e}")
    return {"invite_id": invite_id, "message": f"Invitation sent to {inv_email}"}


@app.get("/api/invitations/{invite_id}")
async def get_invitation(invite_id: str):
    inv = await db.invitations.find_one({"invite_id": invite_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invitation not found")
    group = await db.groups.find_one({"group_id": inv["group_id"]}, {"_id": 0})
    inv["group_name"] = group["name"] if group else "Unknown"
    return inv


@app.post("/api/invitations/{invite_id}/accept")
async def accept_invitation(invite_id: str, request: Request):
    user = await get_current_user(request)
    inv = await db.invitations.find_one({"invite_id": invite_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invitation not found")
    if inv["status"] != "pending":
        raise HTTPException(status_code=400, detail="Invitation already processed")
    if inv["email"] != user["email"]:
        raise HTTPException(status_code=403, detail="This invitation is for a different email")
    existing = await db.group_members.find_one({"group_id": inv["group_id"], "user_id": user["user_id"]}, {"_id": 0})
    if not existing:
        await db.group_members.insert_one({"group_id": inv["group_id"], "user_id": user["user_id"], "role": inv["role"], "joined_at": datetime.now(timezone.utc).isoformat()})
    await db.invitations.update_one({"invite_id": invite_id}, {"$set": {"status": "accepted", "accepted_by_user_id": user["user_id"]}})
    return {"message": "Invitation accepted", "group_id": inv["group_id"]}


@app.get("/api/groups/{group_id}/invitations")
async def list_invitations(group_id: str, request: Request):
    await require_group_role(request, group_id, ["admin"])
    return await db.invitations.find({"group_id": group_id}, {"_id": 0}).sort("created_at", -1).to_list(200)


# ── Funds (Group-Scoped) ───────────────────────────────────────
@app.get("/api/groups/{group_id}/funds")
async def list_funds(group_id: str, request: Request, category: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, search: Optional[str] = None):
    user = await get_current_user(request)
    if not user.get("is_super_admin"):
        r = await get_group_member_role(user["user_id"], group_id)
        if not r: raise HTTPException(status_code=403, detail="Not a member")
    q = {"group_id": group_id}
    if category: q["category"] = category
    if date_from: q.setdefault("date_received", {})["$gte"] = date_from
    if date_to: q.setdefault("date_received", {})["$lte"] = date_to
    if search: q["source_name"] = {"$regex": search, "$options": "i"}
    return await db.funds.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@app.post("/api/groups/{group_id}/funds")
async def create_fund(group_id: str, fund: FundCreate, request: Request):
    user, role = await require_group_role(request, group_id, ["admin", "contributor"])
    fid = f"fund_{uuid.uuid4().hex[:12]}"
    doc = {"fund_id": fid, "group_id": group_id, "source_name": fund.source_name, "amount_inr": fund.amount_inr,
           "category": fund.category, "date_received": fund.date_received, "notes": fund.notes, "attachment_url": fund.attachment_url,
           "added_by_user_id": user["user_id"], "added_by_name": user["name"],
           "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()}
    await db.funds.insert_one(doc)
    return await db.funds.find_one({"fund_id": fid}, {"_id": 0})


@app.put("/api/groups/{group_id}/funds/{fund_id}")
async def update_fund(group_id: str, fund_id: str, fund: FundCreate, request: Request):
    user, role = await require_group_role(request, group_id, ["admin", "contributor"])
    existing = await db.funds.find_one({"fund_id": fund_id, "group_id": group_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Fund not found")
    if role not in ["admin", "super_admin"] and existing["added_by_user_id"] != user["user_id"]: raise HTTPException(status_code=403, detail="Not authorized")
    await db.funds.update_one({"fund_id": fund_id}, {"$set": {"source_name": fund.source_name, "amount_inr": fund.amount_inr, "category": fund.category, "date_received": fund.date_received, "notes": fund.notes, "attachment_url": fund.attachment_url, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return await db.funds.find_one({"fund_id": fund_id}, {"_id": 0})


@app.delete("/api/groups/{group_id}/funds/{fund_id}")
async def delete_fund(group_id: str, fund_id: str, request: Request):
    user, role = await require_group_role(request, group_id, ["admin", "contributor"])
    existing = await db.funds.find_one({"fund_id": fund_id, "group_id": group_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Fund not found")
    if role not in ["admin", "super_admin"] and existing["added_by_user_id"] != user["user_id"]: raise HTTPException(status_code=403, detail="Not authorized")
    await db.funds.delete_one({"fund_id": fund_id})
    return {"message": "Fund deleted"}


# ── Utilizations (Group-Scoped) ─────────────────────────────────
@app.get("/api/groups/{group_id}/utilizations")
async def list_utilizations(group_id: str, request: Request, linked_fund_id: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, spent_by_user_id: Optional[str] = None):
    user = await get_current_user(request)
    if not user.get("is_super_admin"):
        r = await get_group_member_role(user["user_id"], group_id)
        if not r: raise HTTPException(status_code=403, detail="Not a member")
    q = {"group_id": group_id}
    if linked_fund_id: q["linked_fund_id"] = linked_fund_id
    if date_from: q.setdefault("date_spent", {})["$gte"] = date_from
    if date_to: q.setdefault("date_spent", {})["$lte"] = date_to
    if spent_by_user_id: q["spent_by_user_id"] = spent_by_user_id
    return await db.utilizations.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@app.post("/api/groups/{group_id}/utilizations")
async def create_utilization(group_id: str, util: UtilizationCreate, request: Request):
    user, role = await require_group_role(request, group_id, ["admin", "contributor"])
    fund = await db.funds.find_one({"fund_id": util.linked_fund_id, "group_id": group_id}, {"_id": 0})
    if not fund: raise HTTPException(status_code=400, detail="Linked fund not found")
    uid = f"util_{uuid.uuid4().hex[:12]}"
    doc = {"util_id": uid, "group_id": group_id, "purpose": util.purpose, "amount_inr": util.amount_inr, "date_spent": util.date_spent,
           "linked_fund_id": util.linked_fund_id, "linked_fund_name": fund["source_name"],
           "spent_by_user_id": user["user_id"], "spent_by_name": user["name"],
           "notes": util.notes, "receipt_url": util.receipt_url,
           "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()}
    await db.utilizations.insert_one(doc)
    return await db.utilizations.find_one({"util_id": uid}, {"_id": 0})


@app.put("/api/groups/{group_id}/utilizations/{util_id}")
async def update_utilization(group_id: str, util_id: str, util: UtilizationCreate, request: Request):
    user, role = await require_group_role(request, group_id, ["admin", "contributor"])
    existing = await db.utilizations.find_one({"util_id": util_id, "group_id": group_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Utilization not found")
    if role not in ["admin", "super_admin"] and existing["spent_by_user_id"] != user["user_id"]: raise HTTPException(status_code=403, detail="Not authorized")
    fund = await db.funds.find_one({"fund_id": util.linked_fund_id, "group_id": group_id}, {"_id": 0})
    if not fund: raise HTTPException(status_code=400, detail="Linked fund not found")
    await db.utilizations.update_one({"util_id": util_id}, {"$set": {"purpose": util.purpose, "amount_inr": util.amount_inr, "date_spent": util.date_spent, "linked_fund_id": util.linked_fund_id, "linked_fund_name": fund["source_name"], "notes": util.notes, "receipt_url": util.receipt_url, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return await db.utilizations.find_one({"util_id": util_id}, {"_id": 0})


@app.delete("/api/groups/{group_id}/utilizations/{util_id}")
async def delete_utilization(group_id: str, util_id: str, request: Request):
    user, role = await require_group_role(request, group_id, ["admin", "contributor"])
    existing = await db.utilizations.find_one({"util_id": util_id, "group_id": group_id}, {"_id": 0})
    if not existing: raise HTTPException(status_code=404, detail="Utilization not found")
    if role not in ["admin", "super_admin"] and existing["spent_by_user_id"] != user["user_id"]: raise HTTPException(status_code=403, detail="Not authorized")
    await db.utilizations.delete_one({"util_id": util_id})
    return {"message": "Utilization deleted"}


# ── Dashboard (Group-Scoped) ────────────────────────────────────
@app.get("/api/groups/{group_id}/dashboard")
async def get_dashboard(group_id: str, request: Request):
    user = await get_current_user(request)
    if not user.get("is_super_admin"):
        r = await get_group_member_role(user["user_id"], group_id)
        if not r: raise HTTPException(status_code=403, detail="Not a member")
    mf = {"$match": {"group_id": group_id}}
    mu = {"$match": {"group_id": group_id}}
    tf = await db.funds.aggregate([mf, {"$group": {"_id": None, "total": {"$sum": "$amount_inr"}}}]).to_list(1)
    tu = await db.utilizations.aggregate([mu, {"$group": {"_id": None, "total": {"$sum": "$amount_inr"}}}]).to_list(1)
    bc = await db.funds.aggregate([mf, {"$group": {"_id": "$category", "total": {"$sum": "$amount_inr"}}}]).to_list(100)
    mfr = await db.funds.aggregate([mf, {"$group": {"_id": {"$substr": ["$date_received", 0, 7]}, "total": {"$sum": "$amount_inr"}}}, {"$sort": {"_id": 1}}]).to_list(24)
    mur = await db.utilizations.aggregate([mu, {"$group": {"_id": {"$substr": ["$date_spent", 0, 7]}, "total": {"$sum": "$amount_inr"}}}, {"$sort": {"_id": 1}}]).to_list(24)
    tc = tf[0]["total"] if tf else 0
    tu_val = tu[0]["total"] if tu else 0
    bal = tc - tu_val
    pct = round((tu_val / tc * 100), 1) if tc > 0 else 0
    cat = [{"category": c["_id"], "amount": c["total"]} for c in bc]
    ms = set()
    fbm, ubm = {}, {}
    for m in mfr: ms.add(m["_id"]); fbm[m["_id"]] = m["total"]
    for m in mur: ms.add(m["_id"]); ubm[m["_id"]] = m["total"]
    monthly = sorted([{"month": m, "collected": fbm.get(m, 0), "utilized": ubm.get(m, 0)} for m in ms], key=lambda x: x["month"])
    rf = await db.funds.find({"group_id": group_id}, {"_id": 0}).sort("created_at", -1).to_list(5)
    ru = await db.utilizations.find({"group_id": group_id}, {"_id": 0}).sort("created_at", -1).to_list(5)
    activity = []
    for f in rf: activity.append({"type": "fund", "description": f"Fund added: {f['source_name']}", "amount": f["amount_inr"], "date": f["created_at"], "user": f.get("added_by_name", "")})
    for u in ru: activity.append({"type": "utilization", "description": f"Utilized: {u['purpose']}", "amount": u["amount_inr"], "date": u["created_at"], "user": u.get("spent_by_name", "")})
    activity.sort(key=lambda x: x["date"], reverse=True)
    return {"total_collected": tc, "total_utilized": tu_val, "balance": bal, "pct_utilized": pct, "category_breakdown": cat, "monthly_data": monthly, "recent_activity": activity[:10]}


# ── Export (Group-Scoped) ───────────────────────────────────────
@app.get("/api/groups/{group_id}/export/excel")
async def export_excel(group_id: str, request: Request, date_from: Optional[str] = None, date_to: Optional[str] = None, category: Optional[str] = None, user_id: Optional[str] = None):
    user = await get_current_user(request)
    if not user.get("is_super_admin"):
        r = await get_group_member_role(user["user_id"], group_id)
        if not r: raise HTTPException(status_code=403, detail="Not a member")
    fq, uq = {"group_id": group_id}, {"group_id": group_id}
    if category: fq["category"] = category
    if date_from: fq.setdefault("date_received", {})["$gte"] = date_from; uq.setdefault("date_spent", {})["$gte"] = date_from
    if date_to: fq.setdefault("date_received", {})["$lte"] = date_to; uq.setdefault("date_spent", {})["$lte"] = date_to
    if user_id: fq["added_by_user_id"] = user_id; uq["spent_by_user_id"] = user_id
    funds = await db.funds.find(fq, {"_id": 0}).sort("date_received", -1).to_list(5000)
    utils = await db.utilizations.find(uq, {"_id": 0}).sort("date_spent", -1).to_list(5000)
    wb = Workbook()
    hf = Font(bold=True); hfill = PatternFill(start_color="B4D7E8", end_color="B4D7E8", fill_type="solid"); cfmt = '\u20b9#,##0.00'
    ws1 = wb.active; ws1.title = "Funds"
    for ci, h in enumerate(["#","Source","Amount","Category","Date","Added By","Notes"], 1):
        c = ws1.cell(row=1, column=ci, value=h); c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")
    ttf = 0
    for i, f in enumerate(funds, 1):
        ws1.cell(row=i+1, column=1, value=i); ws1.cell(row=i+1, column=2, value=f.get("source_name",""))
        ac = ws1.cell(row=i+1, column=3, value=f.get("amount_inr",0)); ac.number_format = cfmt; ttf += f.get("amount_inr",0)
        ws1.cell(row=i+1, column=4, value=f.get("category","")); ws1.cell(row=i+1, column=5, value=f.get("date_received",""))
        ws1.cell(row=i+1, column=6, value=f.get("added_by_name","")); ws1.cell(row=i+1, column=7, value=f.get("notes",""))
    sr = len(funds)+2; ws1.cell(row=sr, column=2, value="TOTAL").font = Font(bold=True)
    tc = ws1.cell(row=sr, column=3, value=ttf); tc.font = Font(bold=True); tc.number_format = cfmt
    for col in ws1.columns: ws1.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col)+4, 40)

    ws2 = wb.create_sheet("Utilization")
    for ci, h in enumerate(["#","Purpose","Amount","Date","Fund","Spent By","Notes"], 1):
        c = ws2.cell(row=1, column=ci, value=h); c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")
    ttu = 0
    for i, u in enumerate(utils, 1):
        ws2.cell(row=i+1, column=1, value=i); ws2.cell(row=i+1, column=2, value=u.get("purpose",""))
        ac = ws2.cell(row=i+1, column=3, value=u.get("amount_inr",0)); ac.number_format = cfmt; ttu += u.get("amount_inr",0)
        ws2.cell(row=i+1, column=4, value=u.get("date_spent","")); ws2.cell(row=i+1, column=5, value=u.get("linked_fund_name",""))
        ws2.cell(row=i+1, column=6, value=u.get("spent_by_name","")); ws2.cell(row=i+1, column=7, value=u.get("notes",""))
    sr2 = len(utils)+2; ws2.cell(row=sr2, column=2, value="TOTAL").font = Font(bold=True)
    tc2 = ws2.cell(row=sr2, column=3, value=ttu); tc2.font = Font(bold=True); tc2.number_format = cfmt
    for col in ws2.columns: ws2.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col)+4, 40)

    ws3 = wb.create_sheet("Summary")
    for ci, h in enumerate(["Metric","Value"], 1): c = ws3.cell(row=1, column=ci, value=h); c.font = hf; c.fill = hfill
    bal = ttf-ttu; pct = round((ttu/ttf*100),1) if ttf>0 else 0
    rows = [("Total Collected",ttf),("Total Utilized",ttu),("Balance",bal),("% Utilized",f"{pct}%")]
    for cc in await db.funds.aggregate([{"$match":fq},{"$group":{"_id":"$category","total":{"$sum":"$amount_inr"}}}]).to_list(100): rows.append((f"Category: {cc['_id']}",cc['total']))
    for i,(m,v) in enumerate(rows,2):
        ws3.cell(row=i, column=1, value=m); cl = ws3.cell(row=i, column=2, value=v)
        if isinstance(v,(int,float)): cl.number_format = cfmt
    for col in ws3.columns: ws3.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col)+4, 40)
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    g = await db.groups.find_one({"group_id": group_id}, {"_id": 0})
    gn = (g["name"] if g else "report").replace(" ","_")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=fundtrack_{gn}_{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.xlsx"})


@app.get("/api/health")
async def health():
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
